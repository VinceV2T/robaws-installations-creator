#!/usr/bin/env python3
"""
Robaws installations creator.

Wat dit script doet
-------------------
Iedere uitvoering:
1. Leest de masterlijst van installatie-art.nummers in vanuit data/*.xlsx.
2. Haalt purchase-supply-orders (bestelbonnen) op die in de laatste DAYS_BACK
   dagen aangemaakt of gewijzigd zijn.
3. Voor elke bestelbon worden de lijnen geladen (?include=article,order).
4. Voor elke lijn waar het artikel in de masterlijst zit EN waar de
   materieel-kolom (materialId) nog leeg is:
   a. Haalt de gekoppelde sales order op (voor adres, klant, eindklant, project).
   b. Maakt 1 installatie aan per stuk (op basis van line.quantity) via
      POST /api/v2/installations, met serienummer LEEG en
      assignedProjectId = order.projectId. Door die project-link verschijnt
      de installatie automatisch in de orderweergave.
   c. PATCHt de bestelbon-lijn met materialId = nieuwe installatie-id, zodat
      de materieel-kolom op de bestelbon meteen gevuld is en het artikel
      doorheen de volledige flow zichtbaar wordt (voor latere onderhouds-
      opdrachten vanuit het installatie-tabblad).
5. Schrijft een rapport (create_report.json) en stuurt eventueel een mail bij
   echte problemen (geen project, geen leveradres, gekoppelde order ontbreekt,
   aanmaak/PATCH gefaald). Lege serienummers zijn GEEN probleem in deze fase
   en worden NIET in de mail gerapporteerd.

Veiligheid
----------
- DRY_RUN staat default op true. Schakel pas naar false zodra de dry-run
  geverifieerd is op echte data.
- Bij API-fouten wordt de installatie niet aangemaakt - error in 'errors'.
- Idempotent: lijnen waar materialId al gezet is worden overgeslagen. Een
  bestelbon-lijn voor 5 compressoren krijgt dus 5 installaties bij de eerste
  run, en wordt bij volgende runs gewoon overgeslagen.

Env-variabelen
--------------
ROBAWS_API_KEY      verplicht
ROBAWS_API_SECRET   verplicht
DRY_RUN             "true" (default) of "false"
DAYS_BACK           default 14
TEST_BESTELBON      optioneel: enkel deze logicId verwerken
ROBAWS_BASE_URL     default https://app.robaws.com

Mail (optioneel; leeg = mail uit):
SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, MAIL_FROM, MAIL_TO
"""
import datetime as dt
import glob
import json
import os
import smtplib
import sys
from email.mime.text import MIMEText
from pathlib import Path

import requests
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth

# ---------- Configuratie (module-niveau) -----------------------------------
DEFAULT_STATUS = "Actief"
DATA_DIR = Path(__file__).parent / "data"
ARTICLE_NUMBER_COL = "Art.nummer"      # kolomnaam in de Excel(s)


# ---------- Hoofdlogica ----------------------------------------------------
def main() -> int:
    base_url = os.environ.get("ROBAWS_BASE_URL", "https://app.robaws.com").rstrip("/")
    api_key = os.environ.get("ROBAWS_API_KEY", "").strip()
    api_secret = os.environ.get("ROBAWS_API_SECRET", "").strip()
    dry_run = os.environ.get("DRY_RUN", "true").strip().lower() in ("1", "true", "yes", "ja")
    days_back = int(os.environ.get("DAYS_BACK", "14"))
    test_bestelbon = os.environ.get("TEST_BESTELBON", "").strip()

    if not (api_key and api_secret):
        print("FOUT: API-credentials ontbreken (ROBAWS_API_KEY / ROBAWS_API_SECRET).",
              file=sys.stderr)
        return 1

    cutoff = (dt.date.today() - dt.timedelta(days=days_back)).isoformat()
    mode = "DRY-RUN" if dry_run else "LIVE"
    print("=" * 60)
    print(f"Robaws installations creator - modus: {mode}")
    print(f"Host: {base_url}")
    print(f"Cutoff: {cutoff} (laatste {days_back} dagen)")
    if test_bestelbon:
        print(f"TEST_BESTELBON: {test_bestelbon} (enkel deze bestelbon wordt verwerkt)")
    print(f"Tijd: {dt.datetime.now().isoformat(timespec='seconds')}")
    print("=" * 60)

    # 1) Masterlijst inlezen
    master_numbers = load_master_article_numbers(DATA_DIR)
    if not master_numbers:
        print("FOUT: Geen artikelnummers gevonden in data/*.xlsx.", file=sys.stderr)
        return 1
    print(f"Masterlijst: {len(master_numbers)} unieke art.nummers ingeladen.\n")

    session = requests.Session()
    session.auth = HTTPBasicAuth(api_key, api_secret)
    session.headers.update({
        "Accept": "application/json",
        "Content-Type": "application/json",
    })

    report = {
        "dry_run": dry_run,
        "cutoff_date": cutoff,
        "started_at": dt.datetime.now().isoformat(timespec="seconds"),
        "bestelbon_count": 0,
        "lines_evaluated": 0,
        "lines_matched": 0,
        "installations_created": [],         # nieuwe installaties via POST
        "installations_reused": [],          # bestaande installaties hergebruikt (geen POST)
        "lines_linked": [],                  # bestelbon-lijnen met materialId gepatcht
        "skipped_already_linked": [],        # materialId was al gezet
        "skipped_not_in_masterlist": 0,      # tellen, niet detail
        "errors": [],
    }

    # 2) Bestelbonnen ophalen
    try:
        bestelbonnen = fetch_recent_bestelbonnen(session, base_url, cutoff)
    except Exception as exc:
        print(f"FOUT bij ophalen bestelbonnen: {exc}", file=sys.stderr)
        report["errors"].append({"stage": "list_bestelbonnen", "error": str(exc)})
        save_report(report)
        return 2

    if test_bestelbon:
        bestelbonnen = [b for b in bestelbonnen if b.get("logicId") == test_bestelbon]

    report["bestelbon_count"] = len(bestelbonnen)
    print(f"Gevonden bestelbonnen sinds {cutoff}: {len(bestelbonnen)}\n")

    if test_bestelbon and not bestelbonnen:
        print(f"WAARSCHUWING: TEST_BESTELBON={test_bestelbon} niet gevonden in cutoff-window.")

    # Caches
    sales_order_cache = {}     # salesOrderId -> SalesOrderReadDTO
    # Installaties die we tijdens DEZE run reeds geclaimd hebben (om binnen
    # 1 run niet 2x dezelfde te hergebruiken voor 2 verschillende lijnen).
    claimed_installation_ids = set()

    # 3) Per bestelbon de lijnen verwerken
    for bb in bestelbonnen:
        bb_id = bb.get("id")
        bb_logic = bb.get("logicId") or bb_id

        try:
            lines = fetch_bestelbon_lines(session, base_url, bb_id)
        except Exception as exc:
            report["errors"].append({
                "stage": "fetch_lines",
                "bestelbon": bb_logic, "bestelbon_id": bb_id,
                "error": str(exc),
            })
            print(f"  [ERROR] {bb_logic}: kon lijnen niet ophalen: {exc}")
            continue

        for line in lines:
            line_id = line.get("id")
            line_type = line.get("type")
            if line_type == "TEXT":
                continue  # informatieve tekstlijnen overslaan
            report["lines_evaluated"] += 1

            article = line.get("article") or {}
            article_id = line.get("articleId") or article.get("id")
            article_number = article.get("articleNumber") or ""
            article_name = article.get("name") or line.get("description") or ""
            article_brand = article.get("brand") or ""

            # Filter: enkel artikelen uit de masterlijst
            if not article_number or str(article_number) not in master_numbers:
                report["skipped_not_in_masterlist"] += 1
                continue
            report["lines_matched"] += 1

            # Idempotency: lijn waarvan materieel-kolom (materialId) al gezet is
            # is reeds verwerkt in een vorige run.
            if line.get("materialId"):
                report["skipped_already_linked"].append({
                    "bestelbon": bb_logic, "line_id": line_id,
                    "article_number": article_number,
                    "material_id": line.get("materialId"),
                })
                continue

            # Bestelbon-lijn moet gekoppeld zijn aan een sales order
            sales_order_id = line.get("orderId")
            if not sales_order_id:
                report["errors"].append({
                    "stage": "no_order",
                    "bestelbon": bb_logic, "line_id": line_id,
                    "article_number": article_number,
                    "error": "lijn op bestelbon heeft geen orderId (sales order)",
                })
                print(f"  [PROBLEEM] {bb_logic} lijn {line_id}: geen gekoppelde order")
                continue

            # Sales order ophalen (voor adres, klant, eindklant, project)
            if sales_order_id not in sales_order_cache:
                try:
                    sales_order_cache[sales_order_id] = fetch_sales_order(
                        session, base_url, sales_order_id
                    )
                except Exception as exc:
                    sales_order_cache[sales_order_id] = {"_error": str(exc)}
            so = sales_order_cache[sales_order_id]
            if "_error" in so:
                report["errors"].append({
                    "stage": "fetch_sales_order",
                    "bestelbon": bb_logic, "line_id": line_id,
                    "sales_order_id": sales_order_id,
                    "error": so["_error"],
                })
                continue

            project_id = so.get("projectId")
            client_id = so.get("clientId")
            end_client_id = so.get("endClientId")
            address = so.get("address") or {}
            company_id = so.get("companyId")
            subscription_id = so.get("subscriptionId")

            # Adres-validatie: minstens stad/postcode of straat aanwezig?
            if not (address.get("addressLine1") or address.get("city")):
                report["errors"].append({
                    "stage": "no_address",
                    "bestelbon": bb_logic, "line_id": line_id,
                    "sales_order_id": sales_order_id,
                    "error": "gekoppelde order heeft geen leveradres",
                })
                print(f"  [PROBLEEM] {bb_logic} lijn {line_id}: order zonder leveradres")
                continue

            if not project_id:
                report["errors"].append({
                    "stage": "no_project",
                    "bestelbon": bb_logic, "line_id": line_id,
                    "sales_order_id": sales_order_id,
                    "error": "gekoppelde order heeft geen project",
                })
                print(f"  [PROBLEEM] {bb_logic} lijn {line_id}: order zonder project")
                continue

            # Aantal installaties op basis van quantity (afgekapt naar int)
            quantity = int(line.get("quantity") or 0)
            if quantity <= 0:
                quantity = 1  # minimum 1 als veiligheid

            # Eerst checken of er al installaties bestaan voor (project, artikel).
            # Enkel installaties die nog niet aan een bestelbon-lijn gekoppeld
            # zijn (vrij) komen in aanmerking voor hergebruik. Anders zouden
            # twee bestelbon-lijnen voor hetzelfde artikel beide naar dezelfde
            # installatie linken.
            existing_installations = []
            try:
                all_existing = find_installations(
                    session, base_url, project_id, article_id
                )
                for inst in all_existing:
                    inst_id = inst.get("id")
                    if not inst_id:
                        continue
                    if inst_id in claimed_installation_ids:
                        continue   # geclaimd door eerdere lijn in deze run
                    if is_installation_in_use(session, base_url, inst_id):
                        continue   # al gelinkt aan een andere bestelbon-lijn
                    existing_installations.append(inst)
            except Exception as exc:
                report["errors"].append({
                    "stage": "find_installations",
                    "bestelbon": bb_logic, "line_id": line_id,
                    "error": str(exc),
                })
                continue

            # Payload bouwen voor de installatie (enkel nodig als we POST'en)
            payload = {
                "name": article_name or article_number,
                "serialNumber": "",            # bewust leeg, ingevuld bij levering
                "brand": article_brand,
                "articleId": article_id,
                "supplierId": bb.get("supplierId"),
                "subscriptionId": subscription_id,
                "status": DEFAULT_STATUS,
                "installation": True,
                "assignedProjectId": project_id,
                "assignedClientId": client_id,
                "assignedEndClientId": end_client_id,
                "companyId": company_id,
                "address": {
                    "addressLine1": address.get("addressLine1") or "",
                    "addressLine2": address.get("addressLine2") or "",
                    "postalCode": address.get("postalCode") or "",
                    "city": address.get("city") or "",
                    "country": address.get("country") or "",
                },
                "extraFields": {
                    "source_bestelbon": bb_logic,
                    "source_bestelbon_line_id": str(line_id),
                },
            }
            # Lege strings & None weghalen om Robaws niet te confuseren
            payload = {k: v for k, v in payload.items() if v not in (None, "")}

            # Bepaal hoeveel installaties we nog moeten aanmaken.
            # Bestaande installaties hergebruiken we; enkel het tekort wordt
            # aangevuld met nieuwe POST'en.
            already_count = len(existing_installations)
            to_create = max(0, quantity - already_count)

            # Het materialId voor de PATCH komt bij voorkeur uit een
            # bestaande installatie (zo blijft de manueel aangemaakte fiche
            # leidend); anders uit de eerste nieuw aangemaakte installatie.
            first_installation_id = None
            if existing_installations:
                first_installation_id = existing_installations[0].get("id")
                for inst in existing_installations[:quantity]:
                    inst_id = inst.get("id")
                    claimed_installation_ids.add(inst_id)
                    report["installations_reused"].append({
                        "bestelbon": bb_logic, "line_id": line_id,
                        "project_id": project_id,
                        "article_number": article_number,
                        "installation_id": inst_id,
                        "installation_name": inst.get("name"),
                    })
                    print(f"  [HERGEBRUIK] {bb_logic} lijn {line_id}: "
                          f"bestaande installatie {inst_id} ({inst.get('name')}) "
                          f"voor {article_number} gevonden, geen nieuwe gemaakt")

            # Aanvullen met nieuwe installaties indien nodig
            for n in range(to_create):
                if dry_run:
                    report["installations_created"].append({
                        "bestelbon": bb_logic, "line_id": line_id,
                        "project_id": project_id,
                        "article_number": article_number,
                        "article_name": article_name,
                        "action": "WOULD_POST",
                        "payload_preview": {
                            "name": payload.get("name"),
                            "brand": payload.get("brand"),
                            "city": (payload.get("address") or {}).get("city"),
                        },
                    })
                    print(f"  [DRY] {bb_logic} lijn {line_id}: "
                          f"zou installatie {already_count+n+1}/{quantity} aanmaken voor "
                          f"{article_number} ({article_name}) op project {project_id}")
                    if first_installation_id is None and n == 0:
                        first_installation_id = "<DRY_ID>"
                else:
                    try:
                        created = create_installation(session, base_url, payload)
                        installation_id = created.get("id")
                        if installation_id:
                            claimed_installation_ids.add(installation_id)
                        if first_installation_id is None and n == 0:
                            first_installation_id = installation_id
                        report["installations_created"].append({
                            "bestelbon": bb_logic, "line_id": line_id,
                            "project_id": project_id,
                            "article_number": article_number,
                            "article_name": article_name,
                            "action": "POST",
                            "installation_id": installation_id,
                        })
                        print(f"  [LIVE] {bb_logic} lijn {line_id}: "
                              f"installatie {already_count+n+1}/{quantity} aangemaakt "
                              f"(id={installation_id}) voor {article_number}")
                    except Exception as exc:
                        report["errors"].append({
                            "stage": "create_installation",
                            "bestelbon": bb_logic, "line_id": line_id,
                            "article_number": article_number,
                            "error": str(exc),
                        })
                        print(f"  [ERROR] {bb_logic} lijn {line_id}: "
                              f"aanmaak gefaald: {exc}")
                        break  # geen verdere pogingen voor deze lijn

            # Materieel-kolom op de bestelbon-lijn invullen (PATCH).
            if first_installation_id:
                if dry_run:
                    report["lines_linked"].append({
                        "bestelbon": bb_logic, "line_id": line_id,
                        "action": "WOULD_PATCH_MATERIAL",
                        "material_id": first_installation_id,
                    })
                    print(f"  [DRY] {bb_logic} lijn {line_id}: "
                          f"zou materialId zetten op {first_installation_id}")
                else:
                    try:
                        patch_bestelbon_line_material(
                            session, base_url, bb_id, line_id, first_installation_id
                        )
                        report["lines_linked"].append({
                            "bestelbon": bb_logic, "line_id": line_id,
                            "action": "PATCH_MATERIAL",
                            "material_id": first_installation_id,
                        })
                        print(f"  [LIVE] {bb_logic} lijn {line_id}: "
                              f"materialId gezet op {first_installation_id}")
                    except Exception as exc:
                        report["errors"].append({
                            "stage": "patch_bestelbon_line",
                            "bestelbon": bb_logic, "line_id": line_id,
                            "material_id": first_installation_id,
                            "error": str(exc),
                        })
                        print(f"  [ERROR] {bb_logic} lijn {line_id}: "
                              f"PATCH materialId gefaald: {exc}")

    # 4) Samenvatting
    print()
    print("=" * 60)
    print("Samenvatting")
    print("=" * 60)
    print(f"Modus                       : {mode}")
    print(f"Bestelbonnen gescand        : {report['bestelbon_count']}")
    print(f"Lijnen geëvalueerd          : {report['lines_evaluated']}")
    print(f"Lijnen matched (masterlist) : {report['lines_matched']}")
    print(f"Installaties aangemaakt     : {len(report['installations_created'])}")
    print(f"Installaties hergebruikt    : {len(report['installations_reused'])}")
    print(f"Bestelbon-lijnen gelinkt    : {len(report['lines_linked'])}")
    print(f"Skipped (materialId al gezet): {len(report['skipped_already_linked'])}")
    print(f"Skipped (niet in masterlist): {report['skipped_not_in_masterlist']}")
    print(f"Errors                      : {len(report['errors'])}")

    if report["errors"]:
        print("\nProblemen:")
        for e in report["errors"][:20]:
            print(f"  - {e}")

    save_report(report)

    # 5) Mail bij echte problemen
    if report["errors"]:
        try:
            send_problem_mail(report, mode)
        except Exception as exc:
            print(f"  [WAARSCHUWING] kon probleem-mail niet versturen: {exc}",
                  file=sys.stderr)

    # We falen NIET op individuele errors (zoals de invoice linker).
    return 0


# ---------- Masterlijst -----------------------------------------------------
def load_master_article_numbers(data_dir: Path) -> set:
    """Lees alle .xlsx bestanden in data/ in en bouw een set van art.nummers.

    Verwacht een kolom met de naam ARTICLE_NUMBER_COL ('Art.nummer').
    Sheets zonder die kolom worden genegeerd. Lege cellen overgeslagen.
    """
    numbers = set()
    for path in sorted(data_dir.glob("*.xlsx")):
        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as exc:
            print(f"WAARSCHUWING: kan {path.name} niet openen: {exc}", file=sys.stderr)
            continue
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            if not header:
                continue
            try:
                col_idx = [str(c).strip() if c is not None else "" for c in header]\
                    .index(ARTICLE_NUMBER_COL)
            except ValueError:
                continue
            for row in rows:
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None:
                    continue
                # Forceer string (Excel kan ints maken)
                s = str(val).strip()
                if s:
                    numbers.add(s)
        wb.close()
    return numbers


# ---------- Robaws API-helpers ---------------------------------------------
def fetch_recent_bestelbonnen(session, base_url, cutoff_iso):
    """Haal purchase-supply-orders op sinds cutoff_iso.

    Loopt pagina's van laatste naar eerste en filtert client-side op datum,
    identiek patroon als robaws-invoice-linker (dat in productie draait).
    Geen server-side `updatedFrom` of `sort` (die zijn bij sommige Robaws-
    instanties niet beschikbaar en geven dan een 400).
    """
    list_url = f"{base_url}/api/v2/purchase-supply-orders"
    r = session.get(list_url, params={"size": 100, "page": 0}, timeout=30)
    r.raise_for_status()
    payload = r.json()
    total_pages = (
        payload.get("totalPages")
        or (payload.get("page") or {}).get("totalPages")
        or 1
    )
    items = []
    older_streak = 0
    for page in range(total_pages - 1, -1, -1):
        rr = session.get(list_url, params={"size": 100, "page": page}, timeout=30)
        if rr.status_code != 200:
            break
        page_items = rr.json().get("items") or []
        if not page_items:
            break
        page_added = 0
        for bb in page_items:
            bb_date = (
                bb.get("date")
                or (bb.get("updatedAt") or bb.get("createdAt") or "")[:10]
            )
            if bb_date and bb_date >= cutoff_iso:
                items.append(bb)
                page_added += 1
        if page_added == 0:
            older_streak += 1
            if older_streak >= 2:
                break
        else:
            older_streak = 0
    return items


def fetch_bestelbon_lines(session, base_url, bestelbon_id):
    """Haal lijnen van een bestelbon op met artikel-info ge-include.

    Gebruikt een single include-waarde (Robaws ondersteunt geen
    comma-separated include in alle versies). Order-info halen we apart
    via fetch_sales_order zodra we de orderId kennen.
    """
    url = f"{base_url}/api/v2/purchase-supply-orders/{bestelbon_id}/line-items"
    params = {"include": "article", "size": 200, "page": 0}
    lines = []
    while True:
        r = session.get(url, params=params, timeout=30)
        r.raise_for_status()
        payload = r.json()
        page_items = payload.get("items") or []
        lines.extend(page_items)
        total_pages = (
            payload.get("totalPages")
            or (payload.get("page") or {}).get("totalPages")
            or 1
        )
        if params["page"] + 1 >= total_pages or not page_items:
            break
        params["page"] += 1
    return lines


def fetch_sales_order(session, base_url, sales_order_id):
    url = f"{base_url}/api/v2/sales-orders/{sales_order_id}"
    r = session.get(url, timeout=30)
    r.raise_for_status()
    return r.json()


def find_installations(session, base_url, project_id, article_id):
    """Zoek alle installaties die al bestaan voor (projectId, articleId).

    Geeft de lijst items terug. Lege lijst betekent: nog niets aangemaakt.
    """
    url = f"{base_url}/api/v2/installations"
    params = {
        "projectId": project_id,
        "articleId": article_id,
        "size": 100,
        "page": 0,
    }
    r = session.get(url, params=params, timeout=30)
    r.raise_for_status()
    payload = r.json()
    return payload.get("items") or []


def is_installation_in_use(session, base_url, installation_id):
    """True als er al een bestelbon-lijn bestaat met materialId = installation_id.

    Vermijdt dat we eenzelfde installatie aan twee verschillende bestelbon-
    lijnen koppelen.
    """
    if not installation_id:
        return False
    url = f"{base_url}/api/v2/purchase-supply-orders"
    params = {"materialId": installation_id, "size": 1, "page": 0}
    try:
        r = session.get(url, params=params, timeout=30)
    except requests.RequestException:
        # Bij netwerkfout: conservatief als 'in use' beschouwen om geen
        # foute hergebruik te doen.
        return True
    if r.status_code != 200:
        return True
    return bool(r.json().get("items") or [])


def create_installation(session, base_url, payload):
    url = f"{base_url}/api/v2/installations"
    r = session.post(url, json=payload, timeout=30)
    if r.status_code in (200, 201):
        return r.json()
    raise RuntimeError(
        f"POST installation faalde: status={r.status_code}, body={r.text[:300]!r}"
    )


def patch_bestelbon_line_material(session, base_url, bestelbon_id, line_id, material_id):
    """Zet de materieel-kolom (materialId) op een bestelbon-lijn via PATCH.

    PATCH muteert enkel de meegegeven velden, dus we hoeven niet de hele lijn
    over te sturen.
    """
    url = (f"{base_url}/api/v2/purchase-supply-orders/"
           f"{bestelbon_id}/line-items/{line_id}")
    body = {"materialId": material_id}
    r = session.patch(url, json=body, timeout=30)
    if r.status_code in (200, 204):
        return url
    raise RuntimeError(
        f"PATCH materialId faalde op {url}: status={r.status_code}, "
        f"body={r.text[:300]!r}"
    )


# ---------- Rapportage ------------------------------------------------------
def save_report(report):
    report["finished_at"] = dt.datetime.now().isoformat(timespec="seconds")
    with open("create_report.json", "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2, ensure_ascii=False, default=str)
    print(f"\nRapport bewaard als create_report.json")


def send_problem_mail(report, mode):
    host = os.environ.get("SMTP_HOST", "").strip()
    if not host:
        print("  [INFO] SMTP_HOST niet ingesteld, geen mail verstuurd.")
        return
    port = int(os.environ.get("SMTP_PORT", "587"))
    user = os.environ.get("SMTP_USER", "").strip()
    pwd = os.environ.get("SMTP_PASSWORD", "").strip()
    mail_from = os.environ.get("MAIL_FROM", user).strip()
    mail_to = os.environ.get("MAIL_TO", "vincent@v2technics.be").strip()

    lines = [
        f"Robaws installations creator - rapport ({mode})",
        f"Tijdstip: {dt.datetime.now().isoformat(timespec='seconds')}",
        "",
        f"Bestelbonnen gescand: {report.get('bestelbon_count')}",
        f"Installaties aangemaakt: {len(report.get('installations_created') or [])}",
        f"Errors: {len(report.get('errors') or [])}",
        "",
        "Problemen:",
    ]
    for e in (report.get("errors") or [])[:50]:
        lines.append(f"  - [{e.get('stage')}] bestelbon {e.get('bestelbon')} "
                     f"lijn {e.get('line_id')}: {e.get('error')}")
    if len(report.get("errors") or []) > 50:
        lines.append(f"  ... en nog {len(report['errors']) - 50} meer (zie artifact)")

    msg = MIMEText("\n".join(lines), "plain", "utf-8")
    msg["Subject"] = (f"[Robaws] Installaties-creator: "
                      f"{len(report.get('errors') or [])} probleem(en)")
    msg["From"] = mail_from
    msg["To"] = mail_to

    with smtplib.SMTP(host, port, timeout=30) as smtp:
        smtp.starttls()
        if user and pwd:
            smtp.login(user, pwd)
        smtp.sendmail(mail_from, [mail_to], msg.as_string())
    print(f"  [INFO] probleem-mail verzonden naar {mail_to}")


if __name__ == "__main__":
    sys.exit(main())
